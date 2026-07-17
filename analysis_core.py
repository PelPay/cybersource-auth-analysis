"""
CyberSource merchant authorization-response analysis — core logic.

Pure functions, no Streamlit imports. Given a raw CyberSource Transaction Detail
Report (CSV), produce the merchant-level authorization response analysis and a
formatted Excel workbook.

Public entry points:
    load_rows(source)      -> (idx, rows)     source = path str or bytes
    analyze(idx, rows)     -> dict            all computed structures + metrics
    build_workbook(a, out) -> out             writes the .xlsx
    run(source, out)       -> dict            load + analyze + build; returns summary
"""
import csv, io
from collections import defaultdict, Counter, OrderedDict
from datetime import datetime, timedelta
import openpyxl

# Reports are issued with UTC timestamps ('...Z') but the business day is local.
# Lagos is UTC+1, so a report window of 23:00Z -> 23:00Z is exactly one local day.
# Change this single constant if the reporting timezone ever differs.
REPORT_TZ_OFFSET_HOURS = 1
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# The only CyberSource ics_rmsg values that contain internal commas. They are
# protected before comma-splitting so a single response description is never torn
# into multiple positions. Extend this list if new comma-bearing messages appear.
KNOWN_COMMA_MSGS = [
    "Negative CAM, dCVV, iCVV, or CVV results",
    "We encountered a Payer Authentication problem: Transaction Lookup Not Successful, Check Transaction Id",
    "Lost card, pick up (fraud account)",
    "Stolen card, pick up (fraud account)",
    "Pick-up card, special condition(not lost/stolen)",
    "Payment processor error Request has timed out, transaction not sent to Processor",
]
_SENT = "\x01"

REQUIRED_COLS = ["merchant_ref_number", "merchant_id", "ics_applications",
                 "ics_rcode", "ics_rflag", "ics_rmsg"]


_DATE_FMTS = ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
              "%m/%d/%Y %H:%M", "%m/%d/%Y", "%d/%m/%Y %H:%M", "%d/%m/%Y")


def _parse_date(v):
    """Parse a report timestamp to its LOCAL business date; None if unparseable.

    Timestamps ending in 'Z' are UTC and are shifted by REPORT_TZ_OFFSET_HOURS so
    they land on the local business day (a 23:00Z->23:00Z window == one local day).
    """
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    for f in _DATE_FMTS:
        try:
            dt = datetime.strptime(s, f)
        except ValueError:
            continue
        if s.endswith("Z"):
            dt = dt + timedelta(hours=REPORT_TZ_OFFSET_HOURS)
        return dt.date()
    return None


def date_label(date_min, date_max):
    """ISO label for a report's coverage: '2026-07-08' or '2026-07-07_to_2026-07-08'."""
    if not date_min or not date_max:
        return None
    if date_min == date_max:
        return date_min.isoformat()
    return f"{date_min.isoformat()}_to_{date_max.isoformat()}"


def day_range_label(date_min, date_max):
    """Human filename label for a span of days:
        one day        -> 'july_8'
        same month     -> 'july_1-10'
        across months  -> 'july_20-august_5'
        across years   -> 'december_28_2026-january_3_2027'
    """
    if not date_min or not date_max:
        return None
    m1 = date_min.strftime("%B").lower()
    m2 = date_max.strftime("%B").lower()
    if date_min == date_max:
        return f"{m1}_{date_min.day}"
    if date_min.year != date_max.year:
        return f"{m1}_{date_min.day}_{date_min.year}-{m2}_{date_max.day}_{date_max.year}"
    if date_min.month == date_max.month:
        return f"{m1}_{date_min.day}-{date_max.day}"
    return f"{m1}_{date_min.day}-{m2}_{date_max.day}"


def _split_apps(v):
    return [x.strip() for x in v.split(",")] if v else [""]


def _split_plain(v):
    return v.split(",") if v is not None else [""]


def _parse_msg(raw):
    s = raw if raw is not None else ""
    for m in KNOWN_COMMA_MSGS:
        if m in s:
            s = s.replace(m, m.replace(",", _SENT))
    return [p.replace(_SENT, ",") for p in s.split(",")]


def _auto_heal_msg(raw, pattern):
    """Rebuild ics_rmsg into `len(pattern)` segments using a reliable presence
    pattern (from ics_rcode / ics_rflag, which never contain internal commas).

    pattern[i] is True where position i should carry a message. Empty positions
    consume one blank token; each run of message positions consumes the matching
    run of non-empty tokens — and when a single message position spans several
    tokens (a description with internal commas), they're rejoined with commas.
    Returns the healed list, or None if it can't be resolved unambiguously.
    """
    toks = (raw or "").split(",")
    N = len(pattern)
    result, ti, i = [], 0, 0
    while i < N:
        if not pattern[i]:                      # blank position -> one empty token
            if ti < len(toks) and toks[ti].strip() == "":
                result.append(""); ti += 1; i += 1
                continue
            return None
        j = i                                   # length of this run of message positions
        while j < N and pattern[j]:
            j += 1
        L = j - i
        block = []                              # consecutive non-empty tokens
        while ti < len(toks) and toks[ti].strip() != "":
            block.append(toks[ti]); ti += 1
        if len(block) == L:                     # one token per position
            result.extend(block)
        elif L == 1 and block:                  # one message split by internal commas
            result.append(",".join(block))
        else:                                   # adjacent messages + extra commas: ambiguous
            return None
        i = j
    if len(result) != N or any(t.strip() for t in toks[ti:]):
        return None
    return result


def _align_rmsg(raw, rcode, rflag, N):
    """Return ics_rmsg as N aligned segments. Tries the known-message fast path,
    then auto-heals against the rcode / rflag presence pattern. Falls back to the
    naive split (which the caller then flags) only if nothing resolves."""
    fast = _parse_msg(raw)
    if len(fast) == N:
        return fast
    for anchor in (rcode, rflag):
        if len(anchor) == N:
            healed = _auto_heal_msg(raw, [bool(x.strip()) for x in anchor])
            if healed and len(healed) == N:
                return healed
    return fast


def _rows_from_xlsx(data):
    """Read the first worksheet of an .xlsx into a list of string-cell rows."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if c is None else str(c) for c in row])
    wb.close()
    return out


def _rows_from_csv_text(text):
    return [r for r in csv.reader(io.StringIO(text))]


def load_rows(source):
    """Load a CyberSource Transaction Detail Report (CSV or XLSX).

    `source` is a filesystem path (str) or the raw file bytes (from an upload).
    File type is detected by content (XLSX begins with the ZIP signature 'PK'),
    so bytes from either a loose upload or a ZIP member work regardless of name.
    Skips any leading metadata rows and locates the true header row (the one
    containing merchant_ref_number / ics_applications). Returns (idx, rows).
    """
    if isinstance(source, (bytes, bytearray)):
        b = bytes(source)
    else:
        with open(source, "rb") as fh:
            b = fh.read()

    if b[:4] == b"PK\x03\x04":                       # .xlsx / .xlsm (zip container)
        all_rows = _rows_from_xlsx(b)
    elif b[:4] == b"\xd0\xcf\x11\xe0":               # legacy .xls (OLE) — unsupported
        raise ValueError("Legacy .xls files are not supported — "
                         "please save the report as .xlsx or .csv.")
    else:                                            # CSV / plain text
        all_rows = _rows_from_csv_text(b.decode("utf-8-sig", errors="replace"))

    # Identify the real header row (skipping any metadata rows) as the row that
    # contains the most of the expected CyberSource columns.
    low_required = [c.lower() for c in REQUIRED_COLS]
    header_i, best = None, 0
    for i, r in enumerate(all_rows[:20]):
        cells = {str(c).strip().lower() for c in r if str(c).strip()}
        score = sum(1 for c in low_required if c in cells)
        if score > best:
            best, header_i = score, i
    if header_i is None or best < 3:
        raise ValueError("Could not find a header row with the expected CyberSource "
                         "columns (merchant_ref_number, merchant_id, ics_applications, "
                         "ics_rcode, ics_rflag, ics_rmsg).")

    header = all_rows[header_i]
    idx = {h: i for i, h in enumerate(header)}
    missing = [c for c in REQUIRED_COLS if c not in idx]
    if missing:
        raise ValueError("Report is missing required columns: " + ", ".join(missing))

    rows = [r for r in all_rows[header_i + 1:] if r and any(c.strip() for c in r)]
    return idx, rows


def analyze(idx, rows):
    """Run the 5-step analysis. Returns a dict of all computed structures."""
    C_REF, C_MID = idx["merchant_ref_number"], idx["merchant_id"]
    C_APP, C_RC, C_RF, C_RM = (idx["ics_applications"], idx["ics_rcode"],
                               idx["ics_rflag"], idx["ics_rmsg"])

    # Step 1: group raw rows by merchant_ref_number (one group == one customer txn)
    groups = OrderedDict()
    for i, row in enumerate(rows):
        groups.setdefault(row[C_REF], []).append(i)

    # Steps 2-3: keep refs that attempted ics_auth; extract result by position
    auth_records = []
    align_fail = []
    for ref, ridx in groups.items():
        auth_row = next((i for i in ridx if "ics_auth" in _split_apps(rows[i][C_APP])), None)
        if auth_row is None:
            continue
        row = rows[auth_row]
        apps = _split_apps(row[C_APP]); N = len(apps)
        pos = apps.index("ics_auth")
        rcode = _split_plain(row[C_RC]); rflag = _split_plain(row[C_RF])
        rmsg = _align_rmsg(row[C_RM], rcode, rflag, N)
        if not (len(rcode) == N and len(rflag) == N and len(rmsg) == N):
            align_fail.append({"ref": ref, "N": N, "rcode": len(rcode),
                               "rflag": len(rflag), "rmsg": len(rmsg), "raw_rmsg": row[C_RM]})

        def at(lst):
            return (lst[pos] if pos < len(lst) else "").strip()

        auth_records.append({"ref": ref, "merchant_id": row[C_MID],
                             "code": at(rcode), "flag": at(rflag), "desc": at(rmsg)})

    # Step 5: group by merchant_id then by (code, description)
    by_merch = defaultdict(Counter)
    merch_total = Counter()
    for rec in auth_records:
        by_merch[rec["merchant_id"]][(rec["code"], rec["desc"])] += 1
        merch_total[rec["merchant_id"]] += 1

    # Raw Data Summary metrics
    ref_merch = {ref: rows[ridx[0]][C_MID] for ref, ridx in groups.items()}
    ref_has_auth = {ref: any("ics_auth" in _split_apps(rows[i][C_APP]) for i in ridx)
                    for ref, ridx in groups.items()}
    raw_per_merch = Counter(row[C_MID] for row in rows)

    summ = defaultdict(lambda: dict(total_raw=0, uniq=set(), auth_ref=set(), nonauth_ref=set(),
                                    raw_in_auth=0, raw_in_nonauth=0, raw_has_auth=0, raw_no_auth=0))
    for ref, ridx in groups.items():
        d = summ[ref_merch[ref]]
        d["uniq"].add(ref)
        (d["auth_ref"] if ref_has_auth[ref] else d["nonauth_ref"]).add(ref)
        for i in ridx:
            d["total_raw"] += 1
            if ref_has_auth[ref]:
                d["raw_in_auth"] += 1
            else:
                d["raw_in_nonauth"] += 1
            if "ics_auth" in _split_apps(rows[i][C_APP]):
                d["raw_has_auth"] += 1
            else:
                d["raw_no_auth"] += 1

    # Report coverage — derived from the data itself (transaction_date / submit_time_utc)
    dcol = None
    for cand in ("transaction_date", "submit_time_utc"):
        if cand in idx:
            dcol = idx[cand]
            break
    dates = []
    if dcol is not None:
        for row in rows:
            if dcol < len(row):
                d = _parse_date(row[dcol])
                if d:
                    dates.append(d)
    date_min = min(dates) if dates else None
    date_max = max(dates) if dates else None

    order = [m for m, _ in raw_per_merch.most_common()]  # merchants by raw volume desc
    return {
        "auth_records": auth_records, "align_fail": align_fail,
        "by_merch": by_merch, "merch_total": merch_total,
        "summ": summ, "order": order,
        "n_rows": len(rows), "n_unique_ref": len(groups),
        "n_auth": len(auth_records),
        "date_min": date_min, "date_max": date_max,
        "date_label": date_label(date_min, date_max),
    }


def _summary_rows(a, top_n=5):
    """Return (header, body_rows, total_row) for the Raw Data Summary sheet."""
    cols = ["merchant_id", "total_raw_entries", "Unique Transactions", "auth_attempt_transactions",
            "non_auth_transactions", "raw_entries_in_auth_transactions",
            "raw_entries_in_non_auth_transactions", "raw_rows_containing_ics_auth",
            "raw_rows_without_ics_auth", "auth_attempt_transaction_percentage",
            "non_auth_transaction_percentage"]

    def mrow(mid, d):
        uniq, auth, non = len(d["uniq"]), len(d["auth_ref"]), len(d["nonauth_ref"])
        return [mid, d["total_raw"], uniq, auth, non, d["raw_in_auth"], d["raw_in_nonauth"],
                d["raw_has_auth"], d["raw_no_auth"],
                (auth / uniq if uniq else 0), (non / uniq if uniq else 0)]

    body = [mrow(mid, a["summ"][mid]) for mid in a["order"][:top_n]]

    T = Counter()
    for d in a["summ"].values():
        T["total_raw"] += d["total_raw"]; T["uniq"] += len(d["uniq"]); T["auth"] += len(d["auth_ref"])
        T["non"] += len(d["nonauth_ref"]); T["ria"] += d["raw_in_auth"]; T["rina"] += d["raw_in_nonauth"]
        T["rha"] += d["raw_has_auth"]; T["rna"] += d["raw_no_auth"]
    total = ["TOTAL", T["total_raw"], T["uniq"], T["auth"], T["non"], T["ria"], T["rina"],
             T["rha"], T["rna"],
             (T["auth"] / T["uniq"] if T["uniq"] else 0), (T["non"] / T["uniq"] if T["uniq"] else 0)]
    return cols, body, total


def build_workbook(a, out_path, top_n=5):
    HDR = PatternFill("solid", fgColor="1F4E78")
    WHITE = Font(bold=True, color="FFFFFF")
    MHDR = Font(bold=True, size=12)
    PCT = "0.00%"

    def hrow(ws, r, cols):
        for c, v in enumerate(cols, 1):
            cell = ws.cell(r, c, v); cell.fill = HDR; cell.font = WHITE

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # Merchant Response Tables (primary sheet)
    ws2 = wb.create_sheet("Merchant Response Tables")
    r = 1
    for mid in sorted(a["by_merch"].keys()):
        total = a["merch_total"][mid]
        ws2.cell(r, 1, f"Merchant ID: {mid}").font = MHDR; r += 1
        ws2.cell(r, 1, f"Total authorization attempts: {total}").font = Font(bold=True); r += 1
        hrow(ws2, r, ["response_code", "response_description", "number", "percentage_of_total"]); r += 1
        items = sorted(a["by_merch"][mid].items(), key=lambda kv: (-kv[1], kv[0][0], kv[0][1]))
        for (code, desc), n in items:
            ws2.cell(r, 1, code if code else "(none)")
            ws2.cell(r, 2, desc if desc else "(no authorization response returned)")
            ws2.cell(r, 3, n)
            cell = ws2.cell(r, 4, (n / total if total else 0)); cell.number_format = PCT
            r += 1
        r += 1
    for col, w in (("A", 16), ("B", 54), ("C", 10), ("D", 20)):
        ws2.column_dimensions[col].width = w

    # Authorization Attempts (intermediate dataset — one row per auth-attempt ref)
    wsa = wb.create_sheet("Authorization Attempts")
    acols = ["merchant_ref_number", "merchant_id", "response_code",
             "response_flag", "response_description"]
    hrow(wsa, 1, acols)
    ar = 2
    for rec in sorted(a["auth_records"], key=lambda x: (x["merchant_id"], x["ref"])):
        wsa.cell(ar, 1, rec["ref"])
        wsa.cell(ar, 2, rec["merchant_id"])
        wsa.cell(ar, 3, rec["code"] if rec["code"] else "(none)")
        wsa.cell(ar, 4, rec["flag"] if rec["flag"] else "(none)")
        wsa.cell(ar, 5, rec["desc"] if rec["desc"] else "(no authorization response returned)")
        ar += 1
    for col, w in (("A", 34), ("B", 16), ("C", 16), ("D", 20), ("E", 54)):
        wsa.column_dimensions[col].width = w

    # Optional audit summaries — global counts across all authorization attempts
    total_auth = len(a["auth_records"])

    def audit_sheet(title, label, keyfn, empty_label, width_b=54):
        ws_ = wb.create_sheet(title)
        hrow(ws_, 1, [label, "number", "percentage_of_total"])
        counts = Counter(keyfn(rec) or empty_label for rec in a["auth_records"])
        rr = 2
        for key, n in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
            ws_.cell(rr, 1, key); ws_.cell(rr, 2, n)
            cell = ws_.cell(rr, 3, (n / total_auth if total_auth else 0)); cell.number_format = PCT
            rr += 1
        ws_.cell(rr, 1, "TOTAL").font = Font(bold=True)
        ws_.cell(rr, 2, sum(counts.values())).font = Font(bold=True)
        ws_.column_dimensions["A"].width = width_b
        ws_.column_dimensions["B"].width = 10
        ws_.column_dimensions["C"].width = 20

    audit_sheet("Auth Code Summary", "response_code",
                lambda r: r["code"], "(none)", width_b=16)
    audit_sheet("Auth Flag Summary", "response_flag",
                lambda r: r["flag"], "(none)", width_b=24)
    audit_sheet("Auth Message Summary", "response_description",
                lambda r: r["desc"], "(no authorization response returned)", width_b=54)

    # Raw Data Summary
    ws = wb.create_sheet("Raw Data Summary")
    cols, body, total = _summary_rows(a, top_n)
    r = 1; hrow(ws, r, cols); r += 1
    for vals in body:
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            if c in (10, 11):
                cell.number_format = PCT
        r += 1
    for c, v in enumerate(total, 1):
        cell = ws.cell(r, c, v); cell.font = Font(bold=True)
        if c in (10, 11):
            cell.number_format = PCT
    for c in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20 if c == 1 else 16

    # Methodology (instructions embedded in the workbook)
    wsm = wb.create_sheet("Methodology")
    wsm.column_dimensions["A"].width = 110
    from openpyxl.styles import Alignment
    lines = [
        ("CyberSource Merchant Authorization-Response Analysis", MHDR),
        ("", None),
        ("Objective", Font(bold=True, size=12)),
        ("For each merchant, what authorization responses did they receive, how many times did each "
         "occur, and what percentage of that merchant's authorization attempts does each represent?", None),
        ("", None),
        ("Method", Font(bold=True, size=12)),
        ("1. Group raw report rows by merchant_ref_number (one customer transaction each). "
         "request_id is NOT used as the grouping key.", None),
        ("2. Keep only transactions where ics_applications contains 'ics_auth' (order ignored). "
         "One row per unique merchant_ref_number forms the intermediate authorization dataset.", None),
        ("3. Locate the position of 'ics_auth' inside the comma-separated ics_applications list, "
         "then read the authorization result from that same position of ics_rcode (response_code), "
         "ics_rflag (response_flag) and ics_rmsg (response_description).", None),
        ("4. Response descriptions that contain internal commas are protected before splitting so a "
         "single message is never torn across positions. Every row's message count is validated "
         "against ics_rcode.", None),
        ("5. Group the intermediate records by merchant_id, then by response_code + "
         "response_description. number = count; percentage_of_total = number / that merchant's total "
         "authorization attempts.", None),
        ("", None),
        ("Sheets", Font(bold=True, size=12)),
        ("Merchant Response Tables  -  one table per merchant_id (required deliverable).", None),
        ("Raw Data Summary  -  per-merchant reconciliation; first five merchants by raw volume plus a "
         "TOTAL computed from all merchants (required).", None),
        ("Authorization Attempts  -  the intermediate one-row-per-ref dataset (supporting).", None),
        ("", None),
        ("Reconciliation", Font(bold=True, size=12)),
        ("auth_attempt_transactions in Raw Data Summary equals each merchant's Total authorization "
         "attempts in Merchant Response Tables; the TOTAL row equals the number of rows on the "
         "Authorization Attempts sheet.", None),
    ]
    rr = 1
    for text, font in lines:
        cell = wsm.cell(rr, 1, text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1

    wb.save(out_path)
    return out_path


def run(source, out_path, top_n=5):
    idx, rows = load_rows(source)
    a = analyze(idx, rows)
    build_workbook(a, out_path, top_n)
    return {
        "output_path": out_path,
        "rows": a["n_rows"],
        "unique_transactions": a["n_unique_ref"],
        "auth_attempts": a["n_auth"],
        "merchants": sorted(a["by_merch"].keys()),
        "alignment_failures": len(a["align_fail"]),
        "date_min": a["date_min"], "date_max": a["date_max"],
        "date_label": a["date_label"],
    }
